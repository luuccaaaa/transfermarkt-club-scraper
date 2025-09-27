#!/usr/bin/env python3
"""Generate an Excel workbook summarising teams and their players."""
from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent
DEFAULT_CLUBS_DIR = DATA_DIR / "data" / "clubs"
DEFAULT_CLUB_IDS_CSV = DATA_DIR / "data" / "club_ids.csv"
DEFAULT_OUTPUT = DATA_DIR / "data" / "exports" / "team_list.xlsx"
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def load_club_names(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        names = {}
        for row in reader:
            club_id = (row.get("club_id") or "").strip()
            club_name = (row.get("club_name") or "").strip()
            if club_id and club_name:
                names[club_id] = club_name
    return names


def iter_csv_files(directory: Path) -> Iterable[Path]:
    return sorted(p for p in directory.glob("*.csv") if p.is_file())


def read_rows(path: Path) -> Sequence[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        return [dict(row) for row in reader]


def extract_club_id(rows: Sequence[dict[str, str]]) -> str | None:
    for row in rows:
        direct = (row.get("club_id") or "").strip()
        if direct:
            return direct
    for row in rows:
        raw = (row.get("profile_club") or "").strip()
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict):
            club_id = data.get("id")
            if isinstance(club_id, str) and club_id.strip():
                return club_id.strip()
    return None


def infer_club_name(file_path: Path, rows: Sequence[dict[str, str]], club_lookup: Dict[str, str]) -> str:
    club_id = extract_club_id(rows)
    if club_id and club_id in club_lookup:
        return club_lookup[club_id]
    if club_id:
        for row in rows:
            name_hint = (row.get("club_name") or "").strip()
            if name_hint:
                return name_hint
    if club_id:
        # Fall back to any name stored inside the JSON payload.
        for row in rows:
            raw = (row.get("profile_club") or "").strip()
            if not raw:
                continue
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                continue
            name = data.get("name") if isinstance(data, dict) else None
            if isinstance(name, str) and name.strip():
                return name.strip()
    for row in rows:
        name_hint = (row.get("club_name") or "").strip()
        if name_hint:
            return name_hint
    # Use the CSV filename as last resort.
    stem = file_path.stem.replace("_", " ")
    return stem.title()


def sanitise_sheet_name(name: str, used: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub(" ", name)
    cleaned = cleaned.replace("'", "")
    cleaned = cleaned.strip()
    if not cleaned:
        cleaned = "Team"
    cleaned = cleaned[:31]
    base = cleaned
    counter = 1
    while cleaned in used:
        suffix = f"_{counter}"
        cleaned = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        counter += 1
    used.add(cleaned)
    return cleaned


def split_name(full_name: str) -> tuple[str, str]:
    name = full_name.strip()
    if not name:
        return "", ""
    if "," in name:
        family, first = [part.strip() for part in name.split(",", 1)]
        return family, first
    parts = name.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[-1], " ".join(parts[:-1])


def family_name(row: dict[str, str]) -> str:
    family, _ = split_name(row.get("name", ""))
    return family


def first_name(row: dict[str, str]) -> str:
    _, first = split_name(row.get("name", ""))
    return first


def first_non_empty(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None:
            text = str(value).strip()
            if text:
                return text
    return ""


def parse_height_cm(row: dict[str, str]) -> str:
    return first_non_empty(row, "height", "profile_height")


def parse_footedness(row: dict[str, str]) -> str:
    return first_non_empty(row, "foot", "profile_foot")


def parse_market_value(row: dict[str, str]) -> str:
    return first_non_empty(row, "marketValue", "profile_marketValue")


def parse_profile_club(row: dict[str, str]) -> dict[str, str] | None:
    raw = (row.get("profile_club") or "").strip()
    if not raw:
        return None
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return None
    return data if isinstance(data, dict) else None


def parse_joined_on(row: dict[str, str]) -> str:
    direct = first_non_empty(row, "joinedOn")
    if direct:
        return direct
    data = parse_profile_club(row)
    if data:
        joined = data.get("joined")
        if isinstance(joined, str) and joined.strip():
            return joined.strip()
    return ""


def parse_contract_expires(row: dict[str, str]) -> str:
    direct = first_non_empty(row, "contract")
    if direct:
        return direct
    data = parse_profile_club(row)
    if data:
        expires = data.get("contractExpires")
        if isinstance(expires, str) and expires.strip():
            return expires.strip()
    return ""


def parse_signed_from(row: dict[str, str]) -> str:
    return first_non_empty(row, "signedFrom")


def parse_agent(row: dict[str, str]) -> str:
    raw = (row.get("profile_agent") or "").strip()
    if not raw:
        return ""
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return raw
    if isinstance(data, dict):
        name = data.get("name")
        if isinstance(name, str) and name.strip():
            return name.strip()
    return raw


def parse_position(row: dict[str, str]) -> str:
    base = (row.get("position") or "").strip()
    if base:
        return base
    raw = (row.get("profile_position") or "").strip()
    if not raw:
        return ""
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return raw
    if isinstance(data, dict):
        if data.get("main"):
            return str(data["main"])
        other = data.get("other")
        if isinstance(other, list) and other:
            return ", ".join(str(item) for item in other)
    return raw


def parse_number(row: dict[str, str]) -> str:
    number = (row.get("profile_shirtNumber") or row.get("shirtNumber") or "").strip()
    if number.startswith("#"):
        number = number[1:]
    return number


def parse_birthday(row: dict[str, str]) -> str:
    for key in ("profile_dateOfBirth", "dateOfBirth"):
        value = row.get(key)
        if not value:
            continue
        text = value.strip()
        if not text:
            continue
        for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
            except ValueError:
                continue
        return text
    return ""


def parse_nationality(row: dict[str, str]) -> str:
    raw = (row.get("profile_citizenship") or row.get("nationality") or "").strip()
    if not raw:
        return ""
    if raw.startswith("[") and raw.endswith("]"):
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            pass
        else:
            if isinstance(data, list):
                return ", ".join(str(item) for item in data)
    parts = [part.strip() for part in raw.replace(";", ",").split(",") if part.strip()]
    return ", ".join(parts)


def parse_portrait(row: dict[str, str]) -> str:
    for key in ("profile_imageUrl", "imageUrl", "portrait", "image"):
        value = row.get(key)
        if value and value.strip():
            return value.strip()
    return ""


def sort_key_for_row(row: dict[str, str]) -> Tuple[int, str]:
    number = parse_number(row)
    if number:
        digits = ''.join(ch for ch in number if ch.isdigit())
        if digits:
            return int(digits), number
    return float("inf"), number


def autofit_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_length:
                max_length = len(text)
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width


FIELD_DEFINITIONS: Dict[str, tuple[str, Callable[[dict[str, str]], str]]] = {
    "shirt_number": ("Number", parse_number),
    "family_name": ("Family Name", family_name),
    "first_name": ("First Name", first_name),
    "full_name": ("Full Name", lambda row: (row.get("name") or "").strip()),
    "club_name": ("Club", lambda row: (row.get("club_name") or "").strip()),
    "position": ("Position", parse_position),
    "age": ("Age", lambda row: (row.get("age") or "").strip()),
    "birthday": ("Birthday", parse_birthday),
    "nationality": ("Nationality", parse_nationality),
    "height_cm": ("Height (cm)", parse_height_cm),
    "foot": ("Foot", parse_footedness),
    "joined_on": ("Joined", parse_joined_on),
    "signed_from": ("Signed From", parse_signed_from),
    "contract_expires": ("Contract Expires", parse_contract_expires),
    "market_value": ("Market Value", parse_market_value),
    "agent": ("Agent", parse_agent),
    "portrait": ("Portrait Path", parse_portrait),
    "player_id": ("Player ID", lambda row: (row.get("id") or row.get("player_id") or "").strip()),
    "profile_url": ("Profile URL", lambda row: (row.get("profile_url") or "").strip()),
}

DEFAULT_FIELD_ORDER: List[str] = [
    "shirt_number",
    "full_name",
    "position",
    "age",
    "birthday",
    "nationality",
    "height_cm",
    "foot",
    "market_value",
    "joined_on",
    "contract_expires",
]


def get_available_fields() -> Dict[str, str]:
    return {field_id: label for field_id, (label, _) in FIELD_DEFINITIONS.items()}


def resolve_fields(field_ids: Sequence[str] | None) -> List[tuple[str, Callable[[dict[str, str]], str]]]:
    ordered = list(field_ids) if field_ids else list(DEFAULT_FIELD_ORDER)
    resolved: List[tuple[str, Callable[[dict[str, str]], str]]] = []
    for field_id in ordered:
        definition = FIELD_DEFINITIONS.get(field_id)
        if definition:
            resolved.append(definition)
    if not resolved:
        resolved = [FIELD_DEFINITIONS[fid] for fid in DEFAULT_FIELD_ORDER if fid in FIELD_DEFINITIONS]
    return resolved


def build_workbook(
    club_rows: List[tuple[str, Sequence[dict[str, str]]]],
    output_path: Path,
    *,
    field_ids: Sequence[str] | None = None,
) -> None:
    wb = Workbook()
    team_sheet = wb.active
    team_sheet.title = "Team List"
    team_sheet.append(["", "Team"])

    used_sheet_names = {"Team List"}
    resolved_fields = resolve_fields(field_ids)

    for team_name, rows in club_rows:
        team_sheet.append([team_name])

        sheet_name = sanitise_sheet_name(team_name, used_sheet_names)
        ws = wb.create_sheet(title=sheet_name)
        headers = [(label, Font(bold=True)) for label, _ in resolved_fields]
        ws.append(["", *(label for label, _ in headers)])
        for cell, (_, font) in zip(ws[1][1:], headers):
            cell.font = font

        sorted_rows = sorted(rows, key=sort_key_for_row)

        for row in sorted_rows:
            ws.append([
                "",
                *(
                    extractor(row)
                    for _, extractor in resolved_fields
                ),
            ])

        autofit_worksheet(ws)

    autofit_worksheet(team_sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create an Excel workbook summarising team rosters.")
    parser.add_argument("--clubs-dir", type=Path, default=DEFAULT_CLUBS_DIR,
                        help="Directory containing club CSV files (default: clubs)")
    parser.add_argument("--club-ids", type=Path, default=DEFAULT_CLUB_IDS_CSV,
                        help="CSV mapping club IDs to names (default: club_ids.csv)")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help="Path for the generated workbook (default: team_list.xlsx)")
    parser.add_argument("--fields", nargs="+", default=None,
                        help=f"Optional ordered list of field IDs to include (available: {', '.join(sorted(get_available_fields()))})")
    parser.add_argument("--specific-csvs", nargs="+", type=Path, default=None,
                        help="Process only specific CSV files instead of all files in clubs-dir")
    args = parser.parse_args()

    club_lookup = load_club_names(args.club_ids)
    club_rows: List[tuple[str, Sequence[dict[str, str]]]] = []

    # Use specific CSVs if provided, otherwise scan the clubs directory
    if args.specific_csvs:
        files = [path for path in args.specific_csvs if path.exists() and path.suffix.lower() == '.csv']
        if not files:
            print("No valid CSV files found in the specified paths")
            return
    else:
        if not args.clubs_dir.exists():
            raise FileNotFoundError(f"Clubs directory not found: {args.clubs_dir}")
        files = list(iter_csv_files(args.clubs_dir))
        if not files:
            print(f"No CSV files found in {args.clubs_dir}")
            return

    for path in files:
        rows = read_rows(path)
        if not rows:
            continue
        club_name = infer_club_name(path, rows, club_lookup)
        club_rows.append((club_name, rows))

    if not club_rows:
        print("No club data available")
        return

    build_workbook(club_rows, args.output, field_ids=args.fields)
    print(f"Workbook saved to {args.output}")


if __name__ == "__main__":
    main()
